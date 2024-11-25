#SaVannah Hussey
# COSCO 1010
# Lab Section:14
#Sources used: COSC1010_lec15-WorkingWithExcel.pptx.pdf, Online Tutor, 
#https://www.instructables.com/Learn-to-Code-Minecraft-Art-Drawing-With-Microsoft/
# (Took inspiration from the Heart in this website but made some changes)
# ChatGPT "advice for coding pixel art (manually made) into excel with python (no code)" 11/20/2024

# A program that generates pixel art of a simple heart 

# 1st step: insert "pip install openpyxl" into the terminal
# Run code:

import openpyxl
from openpyxl.styles import PatternFill

# Function to set the dimensions of the cells to make them square
def set_square_cells(sheet):
    column_width = 3  
    row_height = 18  

    # Apply column width and row height to all cells
    for col in range(2, 7):  
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = column_width
    for row in range(2, 12):  
        sheet.row_dimensions[row].height = row_height

# Function to fill cells with specific colors
def fill_pixel(sheet, x, y, color):
    cell = sheet.cell(row=y, column=x)
    fill = PatternFill(start_color=color[0:6], end_color=color[0:6], fill_type="solid")
    cell.fill = fill

# Function to generate the pixel art
def generate_pixel_art():
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Set the cells to be square-shaped
    set_square_cells(sheet)


    pixel_data = [
        # Inside the heart (using red)
        (3, 4, "FF0000"), (4, 3, "FF0000"), (5, 3, "FF0000"), (6, 3, "FF0000"), (7, 4, "FF0000"), 
        (8, 5, "FF0000"), (9, 6, "FF0000"), (3, 5, "FF0000"), (4, 4, "FF0000"), (5, 4, "FF0000"),
        (6, 4, "FF0000"), (7, 5, "FF0000"), (8, 6, "FF0000"), (3, 7, "FF0000"), (4, 5, "FF0000"),
        (5, 5, "FF0000"), (6, 5, "FF0000"), (7, 6, "FF0000"), (8, 7, "FF0000"), (3, 8, "FF0000"),
        (4, 6, "FF0000"), (5, 6, "FF0000"), (6, 6, "FF0000"), (7, 7, "FF0000"), (4, 7, "FF0000"),
        (5, 7, "FF0000"), (6, 7, "FF0000"), (7, 8, "FF0000"), (4, 8, "FF0000"), (5, 8, "FF0000"),
        (6, 8, "FF0000"), (4, 9, "FF0000"), (5, 9, "FF0000"), (6, 9, "FF0000"),

        # Black outline (for the heart shape)
        (2, 4, "000000"), (3, 3, "000000"), (4, 2, "000000"), (5, 2, "000000"), (6, 2, "000000"), 
        (7, 3, "000000"), (8, 4, "000000"), (9, 5, "000000"), (2, 5, "000000"), (3, 6, "000000"),
        (4, 10, "000000"), (5, 10, "000000"), (6, 10, "000000"), (7, 9, "000000"), (8, 8, "000000"),
        (9, 7, "000000"), (2, 7, "000000"), (3, 9, "000000"), (2, 8, "000000"), (10, 6, "000000")
    ]

    
    for x, y, color in pixel_data:
        fill_pixel(sheet, x, y, color)

    # Saving the workbook to a file
    wb.save("heart_pixel_art_.xlsx")

generate_pixel_art()







